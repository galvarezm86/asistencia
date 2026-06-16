from flask import Flask, render_template, request, redirect, url_for, session, flash, abort, send_file
import os
import sqlite3
import secrets
import unicodedata
import re
import logging
import io
import qrcode
from functools import wraps
from datetime import timedelta, datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

import psycopg
from psycopg.rows import dict_row


app = Flask(__name__)

CURRENT_DATABASE = os.environ.get(
    "CURRENT_DATABASE",
    "SQLITE"
).upper()

print(
    f"CURRENT_DATABASE = [{CURRENT_DATABASE}]"
)

if CURRENT_DATABASE == "SQLITE":

    DATABASE_URL = os.environ[
        "DATABASE_URL_SQLITE"
    ]

elif CURRENT_DATABASE == "POSTGRESQL":

    DATABASE_URL = os.environ[
        "DATABASE_URL_POSTGRESQL"
    ]

else:

    raise ValueError(
        "Motor de base de datos no soportado"
    )

print(
    f"Base de datos activa: {CURRENT_DATABASE}"
)
    
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
# app.config["SESSION_COOKIE_SECURE"] = True

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

@app.after_request
def agregar_headers_no_cache(response):

    response.headers["Cache-Control"] = (
        "no-store, no-cache, must-revalidate, max-age=0"
    )

    response.headers["Pragma"] = "no-cache"

    response.headers["Expires"] = "0"

    return response
    
ADMIN_USER = os.environ["ADMIN_USER"]
ADMIN_PASSWORD = os.environ["ADMIN_PASSWORD"]
app.secret_key = os.environ["SESSION_SECRET"]

DATABASE_IS_SQLITE = DATABASE_URL.startswith("sqlite:///")
DATABASE_IS_POSTGRESQL = DATABASE_URL.startswith("postgresql://")
if not DATABASE_IS_SQLITE and not DATABASE_IS_POSTGRESQL:
    raise ValueError("Motor de base de datos no soportado")

DB_PARAM = "?" if DATABASE_IS_SQLITE else "%s"

DB_CURRENT_DATE = (
    "DATE('now')"
    if DATABASE_IS_SQLITE
    else "CURRENT_DATE"
)

DATABASE_ERRORS = (
    sqlite3.Error,
    psycopg.Error
)

INTEGRITY_ERRORS = (
    sqlite3.IntegrityError,
    psycopg.IntegrityError
)

    

def get_db_connection():

    if DATABASE_IS_SQLITE:
        return get_sqlite_connection()
    
    if DATABASE_IS_POSTGRESQL:
        return get_postgresql_connection()
    
    raise ValueError("Motor no soportado")


def get_sqlite_connection():
    db_name = DATABASE_URL.replace("sqlite:///", "")
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DATABASE_PATH = os.path.join(BASE_DIR, db_name)
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def get_postgresql_connection():
    conn = psycopg.connect(
        DATABASE_URL,
        row_factory=dict_row # type: ignore[arg-type]
    )
    return conn


def asegurar_csrf_token():

    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)

def login_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):
    
        if not session.get("admin"):
            return redirect(url_for("login"))
    
        return f(*args, **kwargs)
    
    return decorated_function

def validar_csrf():

    token_form = request.form.get("csrf_token")
    
    token_session = session.get("csrf_token")
    
    if (
        not token_form or
        not token_session or
        token_form != token_session
    ):
        abort(403)

MAX_NOMBRE_LENGTH = 100
MAX_CORREO_LENGTH = 255

def normalizar_nombre(nombre):
    
    if nombre is None:
        return ""
    
    nombre = nombre.strip()
    
    nombre = re.sub(
        r"\s+",
        " ",
        nombre
    )
    
    if len(nombre) > MAX_NOMBRE_LENGTH:
        return ""
    
    if not re.fullmatch(
        r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ\s]+",
        nombre
    ):
        return ""
    
    nombre = unicodedata.normalize(
        "NFD",
        nombre
    )
    
    nombre = "".join(
        c for c in nombre
        if unicodedata.category(c) != "Mn"
    )
    
    nombre = nombre.upper()
    
    return nombre

@app.route("/")
def inicio():
    return render_template("inicio.html")

@app.route("/login", methods=["GET", "POST"])
def login():

    asegurar_csrf_token()
    
    if session.get("admin"):
        return redirect(url_for("admin"))
        
    if request.method == "POST":

        validar_csrf()
        
        username = request.form.get("username", "")
        password = request.form.get("password", "")

        if not username or not password:

            flash(
                "Debe ingresar usuario y contraseña",
                "error"
            )

            return redirect(url_for("login"))

        if (
            username == ADMIN_USER and
            password == ADMIN_PASSWORD
        ):

            session.clear()
            session.permanent = True
            session["admin"] = True

            session["csrf_token"] = secrets.token_hex(32)

            flash(
                "Sesión iniciada correctamente",
                "success"
            )

            return redirect(url_for("admin"))

        flash(
            "Credenciales incorrectas",
            "error"
        )

        return redirect(url_for("login"))

    return render_template("login.html")
    
@app.route("/logout", methods=["POST"])
@login_required
def logout():

    validar_csrf()
    
    session.clear()

    flash(
        "Sesión cerrada correctamente",
        "success"
    )

    return redirect(url_for("login"))
    
@app.route("/admin")
@login_required
def admin():

    conn = get_db_connection()

    config = conn.execute(
        """
        SELECT correo_reportes, token_actual
        FROM configuracion
        WHERE id = 1
        """
    ).fetchone()

    if config is None:
        conn.close()
        abort(500)
        
    conn.close()

    return render_template(
        "admin/admin.html",
        correo=config["correo_reportes"],
        token=config["token_actual"]
    )

@app.route("/qr")
@login_required
def qr():

    conn = None

    try:

        conn = get_db_connection()

        config = conn.execute(
            """
            SELECT token_actual
            FROM configuracion
            WHERE id = 1
            """
        ).fetchone()

        if config is None:

            abort(500)

        return render_template(
            "admin/qr.html",
            token=config["token_actual"]
        )

    except DATABASE_ERRORS:

        app.logger.exception(
            "Error en la base de datos al cargar gestión QR"
        )

        flash(
            "Ocurrió un error interno",
            "error"
        )

        return redirect(url_for("admin"))

    finally:

        if conn:
            conn.close()

@app.route("/admin/qr_imagen")
@login_required
def qr_imagen():

    conn = None

    try:

        conn = get_db_connection()

        config = conn.execute(
            """
            SELECT token_actual
            FROM configuracion
            WHERE id = 1
            """
        ).fetchone()

        if config is None:
            abort(500)

        url_formulario = url_for(
            "formulario",
            token=config["token_actual"],
            _external=True
        )

        qr = qrcode.make(url_formulario)

        buffer = io.BytesIO()

        qr.save(buffer, format="PNG")

        buffer.seek(0)

        return send_file(
            buffer,
            mimetype="image/png"
        )

    except DATABASE_ERRORS:

        app.logger.exception(
            "Error en la base de datos al generar QR"
        )

        abort(500)

    finally:

        if conn:
            conn.close()
            
@app.route("/admin/asistencia")
@login_required
def asistencia_admin():

    desde = request.args.get("desde")
    hasta = request.args.get("hasta")

    if desde and hasta and desde > hasta:
        flash(
            "La fecha 'Desde' no puede ser mayor que 'Hasta'",
            "error"
        )
        return redirect(url_for("asistencia_admin"))
        
    where_clauses = []
    params = []

    if desde:

        where_clauses.append(
            f"DATE(fecha_hora) >= {DB_PARAM}"
        )

        params.append(desde)

    if hasta:

        where_clauses.append(
            f"DATE(fecha_hora) <= {DB_PARAM}"
        )

        params.append(hasta)

    where_sql = ""

    if where_clauses:

        where_sql = (
            "WHERE "
            + " AND ".join(where_clauses)
        )

    fechas_sql = f"""
    SELECT DISTINCT
        date(fecha_hora) AS fecha
    FROM asistencias
    {where_sql}
    ORDER BY fecha
    """

    asistencias_sql = f"""
    SELECT
        persona_id,
        date(fecha_hora) AS fecha
    FROM asistencias
    {where_sql}
    """

    personas_sql = f"""
    SELECT DISTINCT
        p.id,
        p.nombre
    FROM personas p
    INNER JOIN asistencias a
        ON a.persona_id = p.id
    """

    personas_params = []

    personas_where_clauses = []

    if desde:

        personas_where_clauses.append(
            f"DATE(a.fecha_hora) >= {DB_PARAM}"
        )

        personas_params.append(desde)

    if hasta:

        personas_where_clauses.append(
            f"DATE(a.fecha_hora) <= {DB_PARAM}"
        )

        personas_params.append(hasta)

    if personas_where_clauses:

        personas_sql += (
            " WHERE "
            + " AND ".join(personas_where_clauses)
        )

    personas_sql += """
     ORDER BY p.nombre
    """

    conn = get_db_connection()

    try:

        fechas = conn.execute(
            fechas_sql,
            params
        ).fetchall()

        personas = conn.execute(
            personas_sql,
            personas_params
        ).fetchall()

        asistencias = conn.execute(
            asistencias_sql,
            params
        ).fetchall()

        # Convertir asistencias a un set para búsquedas rápidas

        if not fechas:
            flash("No existen asistencias para el período seleccionado", "info")
            fechas = []
            datos = []

            return render_template("admin/asistencia.html", fechas=fechas, datos=datos)

        
        asistencias_set = {
            (
                asistencia["persona_id"],
                asistencia["fecha"]
            )
            for asistencia in asistencias
        }

        
        tabla = []

        for persona in personas:

            fila = {
                "nombre": persona["nombre"],
                "total": 0,
                "asistencias": []
            }

            for fecha in fechas:

                presente = (
                    persona["id"],
                    fecha["fecha"]
                ) in asistencias_set

                fila["asistencias"].append(presente)

                if presente:
                    fila["total"] += 1

            
            tabla.append(fila)

        desde_mostrar = None
        hasta_mostrar = None

        if desde:
            desde_mostrar = datetime.strptime(
                desde,
                "%Y-%m-%d"
            ).strftime("%d/%m/%y")

        if hasta:
            hasta_mostrar = datetime.strptime(
                hasta,
                "%Y-%m-%d"
            ).strftime("%d/%m/%y")

        fechas_mostrar = []

        for fecha in fechas:

            valor = fecha["fecha"]

            if isinstance(valor, str):

                fechas_mostrar.append(
                    datetime.strptime(
                        valor,
                        "%Y-%m-%d"
                    ).strftime("%d/%m/%y")
                )

            else:

                fechas_mostrar.append(
                    valor.strftime("%d/%m/%y")
                )
                
        return render_template(
            "admin/asistencia.html",
            fechas=fechas,
            fechas_mostrar=fechas_mostrar,
            tabla=tabla,
            desde=desde,
            hasta=hasta,
            desde_mostrar=desde_mostrar,
            hasta_mostrar=hasta_mostrar
        )

    finally:

        conn.close()

@app.route("/admin/asistencia/excel")
@login_required
def exportar_asistencia_excel():
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")

    if desde and hasta and desde > hasta:
        flash(
            "La fecha 'Desde' no puede ser mayor que 'Hasta'",
            "error"
        )
        return redirect(url_for("asistencia_admin"))

    where_clauses = []
    params = []
    
    if desde:
    
        where_clauses.append(
            f"date(fecha_hora) >= {DB_PARAM}"
        )
    
        params.append(desde)
    
    if hasta:
    
        where_clauses.append(
            f"date(fecha_hora) <= {DB_PARAM}"
        )
    
        params.append(hasta)
    
    where_sql = ""
    
    if where_clauses:
    
        where_sql = (
            "WHERE "
            + " AND ".join(where_clauses)
        )
    
    fechas_sql = f"""
    SELECT DISTINCT
        date(fecha_hora) AS fecha
    FROM asistencias
    {where_sql}
    ORDER BY fecha
    """
    
    asistencias_sql = f"""
    SELECT
        persona_id,
        date(fecha_hora) AS fecha
    FROM asistencias
    {where_sql}
    """
    
    personas_sql = f"""
    SELECT DISTINCT
        p.id,
        p.nombre
    FROM personas p
    INNER JOIN asistencias a
        ON a.persona_id = p.id
    """
    
    personas_params = []
    
    personas_where_clauses = []
    
    if desde:
    
        personas_where_clauses.append(
            f"date(a.fecha_hora) >= {DB_PARAM}"
        )
    
        personas_params.append(desde)
    
    if hasta:
    
        personas_where_clauses.append(
            f"date(a.fecha_hora) <= {DB_PARAM}"
        )
    
        personas_params.append(hasta)
    
    if personas_where_clauses:
    
        personas_sql += (
            " WHERE "
            + " AND ".join(personas_where_clauses)
        )
    
    personas_sql += """
     ORDER BY p.nombre
    """
    
    conn = get_db_connection()
    
    try:
    
        fechas = conn.execute(
            fechas_sql,
            params
        ).fetchall()
    
        personas = conn.execute(
            personas_sql,
            personas_params
        ).fetchall()
    
        asistencias = conn.execute(
            asistencias_sql,
            params
        ).fetchall()
    
        # Convertir asistencias a un set para búsquedas rápidas

        if not fechas:

            flash(
                "No existen asistencias para el período seleccionado",
                "info"
            )
    
            return redirect(
                url_for("asistencia_admin")
            )

        
        asistencias_set = {
            (
                asistencia["persona_id"],
                asistencia["fecha"]
            )
            for asistencia in asistencias
        }
    
        tabla = []
    
        for persona in personas:
    
            fila = {
                "nombre": persona["nombre"],
                "total": 0,
                "asistencias": []
            }
    
            for fecha in fechas:
    
                presente = (
                    persona["id"],
                    fecha["fecha"]
                ) in asistencias_set
    
                fila["asistencias"].append(presente)
    
                if presente:
                    fila["total"] += 1
    
    
            tabla.append(fila)

        wb = Workbook()
        ws = wb.active

        ws.title = "Asistencia"

        encabezados = ["Persona"]

        for fecha in fechas:

            valor = fecha["fecha"]
    
            if isinstance(valor, str):
    
                fecha_formateada = datetime.strptime(
                    valor,
                    "%Y-%m-%d"
                ).strftime("%d/%m/%y")
    
            else:
    
                fecha_formateada = valor.strftime(
                    "%d/%m/%y"
                )
    
            encabezados.append(
                fecha_formateada
            )

        encabezados.append("Total")

        ws.append(encabezados)

        for fila in tabla:

            fila_excel = [fila["nombre"]]
    
            for presente in fila["asistencias"]:
    
                fila_excel.append(
                    "Sí" if presente else "No"
                )
    
            fila_excel.append(
                fila["total"]
            )
    
            ws.append(fila_excel)

        

        for column in ws.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    cell_length = len(
                        str(cell.value)
                    )

                    if cell_length > max_length:
                        max_length = cell_length

                except Exception:
                    pass

            adjusted_width = max_length + 5

            ws.column_dimensions[
                column_letter
            ].width = adjusted_width

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in ws.iter_rows():

            for cell in row:
    
                if cell.column != 1 and cell.value is not None:
    
                    cell.alignment = Alignment(
                        horizontal="center"
                    )
                
        ws.freeze_panes = "B2"
            
        output = io.BytesIO()

        
        wb.save(output)

        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="asistencia.xlsx",
            mimetype=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )
    
    finally:
    
        conn.close()


@app.route("/admin/regenerar_token", methods = ["POST"])
@login_required
def regenerar_token():

    validar_csrf()
    
    nuevo_token = secrets.token_urlsafe(16)

    conn = None

    try:

        conn = get_db_connection()

        conn.execute(
            f"""
            UPDATE configuracion
            SET token_actual = {DB_PARAM}
            WHERE id = 1
            """,
            (nuevo_token,)
        )

        conn.commit()

        flash(
            "QR regenerado correctamente",
            "success"
        )

    except DATABASE_ERRORS:

        if conn:
            conn.rollback()

        app.logger.exception(
            "Error en la base de datos al regenerar token"
        )

        flash(
            "Ocurrió un error interno",
            "error"
        )

    finally:

        if conn:
            conn.close()

    return redirect(url_for("qr"))

@app.route("/admin/qr_pdf", methods=["POST"])
@login_required
def qr_pdf():

    validar_csrf()
    
    # 1. Obtener token actual
    conn = get_db_connection()
    config = conn.execute("SELECT token_actual FROM configuracion WHERE id = 1").fetchone()
    conn.close()

    token = config["token_actual"]

    # 2. URL del formulario
    url_formulario = url_for("formulario", token=token, _external=True)

    # 3. Generar QR
    qr_img = qrcode.make(url_formulario)

    # 4. Crear buffer para PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # 5. Título
    MESES = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre"
    }
    
    hoy = datetime.now()

    dias_hasta_sabado = (5 - hoy.weekday()) % 7

    proximo_sabado = hoy + timedelta(days=dias_hasta_sabado)

    fecha_str = (
        f"{proximo_sabado.day} de "
        f"{MESES[proximo_sabado.month]} de "
        f"{proximo_sabado.year}"
    )
    
    c.setFont("Helvetica-Bold", 28)
    c.drawCentredString(width/2, height - 80, "Asistencia Sábado")
    c.drawCentredString(width/2, height - 120, fecha_str)

    # 6. Insertar QR
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    # Ajustar tamaño QR para ocupar ancho de la página (dejando márgenes)
    qr_width = width - 100  # 50pt margen a cada lado
    qr_height = qr_width
    qr_buffer.seek(0)

    qr_img_reader = ImageReader(qr_buffer)
    
    qr_size = width - 50  # deja márgenes laterales

    x = (width - qr_size) / 2
    y = height - qr_size - 200  # 👈 baja el QR claramente

    c.drawImage(
        qr_img_reader,
        x,
        y,
        width=qr_size,
        height=qr_size
    )

    # 7. Finalizar PDF
    c.showPage()
    c.save()
    buffer.seek(0)

    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name="qr_asistencia.pdf")

@app.route("/admin/configuracion")
@login_required
def configuracion():

    conn = None

    try:

        conn = get_db_connection()

        config = conn.execute(
            """
            SELECT *
            FROM configuracion
            WHERE id = 1
            """
        ).fetchone()

        if not config:
            abort(500)

        return render_template(
            "admin/configuracion.html",
            config=config
        )

    except DATABASE_ERRORS:

        app.logger.exception(
            "Error al cargar configuración"
        )

        flash(
            "Ocurrió un error interno",
            "error"
        )

        return redirect(
            url_for("admin")
        )

    finally:

        if conn:
            conn.close()
    
@app.route("/admin/configuracion/correo", methods=["POST"])
@login_required
def editar_correo():

    validar_csrf()
    
    correo = request.form.get("correo", "").strip()

    if not correo:
        flash("Correo obligatorio", "error")
        return redirect(url_for("admin"))

    if len(correo) > MAX_CORREO_LENGTH:
        flash("Correo demasiado largo", "error")
        return redirect(url_for("admin"))

    patron = r"^[^@]+@[^@]+\.[^@]+$"

    if not re.match(patron, correo):
        flash("Correo inválido", "error")
        return redirect(url_for("admin"))

    conn = None

    try:

        conn = get_db_connection()

        config = conn.execute(
            """
            SELECT correo_reportes
            FROM configuracion
            WHERE id = 1
            """
        ).fetchone()

        if not config:
            abort(500)

        if correo == config["correo_reportes"]:

            flash(
                "No se realizaron cambios",
                "warning"
            )

            return redirect(url_for("admin"))

        conn.execute(
            f"""
            UPDATE configuracion
            SET correo_reportes = {DB_PARAM}
            WHERE id = 1
            """,
            (correo,)
        )

        conn.commit()

        flash(
            "Correo actualizado correctamente",
            "success"
        )

    except DATABASE_ERRORS:

        if conn:
            conn.rollback()

        app.logger.exception(
            "Error en la base de datos al actualizar correo"
        )

        flash(
            "Ocurrió un error interno",
            "error"
        )

    finally:

        if conn:
            conn.close()

    return redirect(url_for("configuracion"))
    
@app.route("/admin/personas")
@login_required
def personas():

    conn = get_db_connection()

    personas = conn.execute(
        """
        SELECT id, nombre
        FROM personas
        WHERE activo = 1
        ORDER BY nombre
        """
    ).fetchall()

    conn.close()

    return render_template(
        "admin/personas.html",
        personas=personas
    )

@app.route("/admin/personas/inactivas")
@login_required
def personas_inactivas():

    conn = get_db_connection()

    personas = conn.execute(
        """
        SELECT id, nombre
        FROM personas
        WHERE activo = 0
        ORDER BY nombre
        """
    ).fetchall()

    conn.close()

    return render_template(
        "admin/personas_inactivas.html",
        personas=personas
    )
    
@app.route("/admin/personas/agregar", methods=["POST"])
@login_required
def agregar_persona():

    validar_csrf()
    
    nombre = request.form.get("nombre", "")

    nombre = normalizar_nombre(nombre)

    if not nombre:
        flash("Nombre inválido", "error")
        return redirect(url_for("personas"))

    conn = None

    try:

        conn = get_db_connection()

        existe = conn.execute(
            f"""
            SELECT 1
            FROM personas
            WHERE nombre = {DB_PARAM}
            """,
            (nombre,)
        ).fetchone()

        if existe:

            flash(
                "Ya existe una persona con ese nombre",
                "error"
            )

            return redirect(url_for("personas"))

        conn.execute(
            f"""
            INSERT INTO personas (nombre)
            VALUES ({DB_PARAM})
            """,
            (nombre,)
        )

        conn.commit()

        flash(
            "Persona agregada correctamente",
            "success"
        )

    except DATABASE_ERRORS:

        if conn:
            conn.rollback()

        app.logger.exception(
            "Error en la base de datos al agregar persona"
        )

        flash(
            "Ocurrió un error interno",
            "error"
        )

    finally:

        if conn:
            conn.close()

    return redirect(url_for("personas"))

@app.route("/admin/personas/desactivar/<int:id>", methods=["POST"])
@login_required
def desactivar_persona(id):

    validar_csrf()
    
    conn = None

    try:

        conn = get_db_connection()

        persona = conn.execute(
            f"""
            SELECT id, activo
            FROM personas
            WHERE id = {DB_PARAM}
            """,
            (id,)
        ).fetchone()

        if not persona:
            abort(404)

        if persona["activo"] == 0:

            flash(
                "La persona ya está inactiva",
                "warning"
            )

            return redirect(url_for("personas"))

        conn.execute(
            f"""
            UPDATE personas
            SET activo = 0
            WHERE id = {DB_PARAM}
            """,
            (id,)
        )

        conn.commit()

        flash(
            "Persona desactivada correctamente",
            "success"
        )

    except DATABASE_ERRORS:

        if conn:
            conn.rollback()

        app.logger.exception(
            "Error en la base de datos al desactivar persona"
        )

        flash(
            "Ocurrió un error interno",
            "error"
        )

    finally:

        if conn:
            conn.close()

    return redirect(url_for("personas"))

@app.route("/admin/personas/reactivar/<int:id>", methods=["POST"])
@login_required
def reactivar_persona(id):

    validar_csrf()
    
    conn = None

    try:

        conn = get_db_connection()

        persona = conn.execute(
            f"""
            SELECT id, activo
            FROM personas
            WHERE id = {DB_PARAM}
            """,
            (id,)
        ).fetchone()

        if not persona:
            abort(404)

        if persona["activo"] == 1:

            flash(
                "La persona ya está activa",
                "warning"
            )

            return redirect(url_for("personas_inactivas"))

        conn.execute(
            f"""
            UPDATE personas
            SET activo = 1
            WHERE id = {DB_PARAM}
            """,
            (id,)
        )

        conn.commit()

        flash(
            "Persona reactivada correctamente",
            "success"
        )

    except DATABASE_ERRORS:

        if conn:
            conn.rollback()

        app.logger.exception(
            "Error en la base de datos al reactivar persona"
        )

        flash(
            "Ocurrió un error interno",
            "error"
        )

    finally:

        if conn:
            conn.close()

    return redirect(url_for("personas"))

@app.route("/admin/persona/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_persona(id):

    conn = None

    try:

        conn = get_db_connection()

        persona = conn.execute(
            f"""
            SELECT id, nombre, activo
            FROM personas
            WHERE id = {DB_PARAM}
            """,
            (id,)
        ).fetchone()

        if not persona:
            abort(404)

        # Obtener URL de origen (next) del query param, fallback según activo
        volver_url = request.args.get("next")
        if not volver_url:
            volver_url = url_for("personas") if persona["activo"] == 1 else url_for("personas_inactivas")

        if request.method == "POST":

            validar_csrf()

            nombre = request.form.get("nombre", "")
            nombre = normalizar_nombre(nombre)

            if not nombre:
                flash("Nombre inválido", "error")
                return redirect(url_for("editar_persona", id=id, next=volver_url))

            if nombre == persona["nombre"]:
                flash("No se realizaron cambios", "warning")
                return redirect(url_for("editar_persona", id=id, next=volver_url))

            existe = conn.execute(
                f"""
                SELECT 1
                FROM personas
                WHERE nombre = {DB_PARAM}
                AND id != {DB_PARAM}
                """,
                (nombre, id)
            ).fetchone()

            if existe:
                flash("Ya existe una persona con ese nombre", "error")
                return redirect(url_for("editar_persona", id=id, next=volver_url))

            conn.execute(
                f"""
                UPDATE personas
                SET nombre = {DB_PARAM}
                WHERE id = {DB_PARAM}
                """,
                (nombre, id)
            )

            conn.commit()
            flash("Nombre actualizado correctamente", "success")

            # Redirigir al origen o fallback
            return redirect(volver_url)

        return render_template(
            "admin/editar_persona.html",
            persona=persona,
            volver_url=volver_url  # enviar al template para el botón "Volver"
        )

    except DATABASE_ERRORS:

        if conn:
            conn.rollback()

        app.logger.exception("Error en la base de datos al editar persona")
        flash("Ocurrió un error interno", "error")
        return redirect(url_for("personas"))

    finally:
        if conn:
            conn.close()
            
@app.route("/formulario/<token>", methods=["GET", "POST"])
def formulario(token):
    
    asegurar_csrf_token()
    
    conn = None

    try:

        conn = get_db_connection()

        config = conn.execute(
            """
            SELECT token_actual
            FROM configuracion
            WHERE id = 1
            """
        ).fetchone()

        if config is None:
            abort(500)

        token_valido = config["token_actual"]

        if token != token_valido:
            abort(404)

        if request.method == "POST":

            validar_csrf()

            persona_id = request.form.get(
                "persona_id",
                ""
            ).strip()

            if not persona_id.isdigit():

                flash(
                    "Persona inválida",
                    "error"
                )

                return redirect(
                    url_for(
                        "formulario",
                        token=token
                    )
                )

            persona_id = int(persona_id)

            if persona_id <= 0:

                flash(
                    "Debe seleccionar una persona",
                    "warning"
                )

                return redirect(
                    url_for(
                        "formulario",
                        token=token
                    )
                )

            persona = conn.execute(
                f"""
                SELECT id, nombre
                FROM personas
                WHERE id = {DB_PARAM}
                AND activo = 1
                """,
                (persona_id,)
            ).fetchone()

            if persona is None:
                abort(404)
            
            existe_sql = f"""
                SELECT 1
                FROM asistencias
                WHERE persona_id = {DB_PARAM}
                AND DATE(fecha_hora) = {DB_CURRENT_DATE}
            """
            
            existe = conn.execute(
                existe_sql,
                (persona_id,)
            ).fetchone()

            if existe:

                flash(
                    "Asistencia ya registrada hoy",
                    "warning"
                )

                return redirect(
                    url_for(
                        "formulario",
                        token=token
                    )
                )

            conn.execute(
                f"""
                INSERT INTO asistencias (persona_id)
                VALUES ({DB_PARAM})
                """,
                (persona_id,)
            )

            conn.commit()

            session["ultimo_nombre"] = persona["nombre"]

            session["ultimo_token"] = token

            return redirect(
                url_for("confirmacion")
            )

        personas = conn.execute(
            """
            SELECT id, nombre
            FROM personas
            WHERE activo = 1
            ORDER BY nombre
            """
        ).fetchall()

        return render_template(
            "formulario.html",
            personas=personas
        )
        
    except INTEGRITY_ERRORS:

        if conn:
            conn.rollback()
    
        flash(
            "Asistencia ya registrada hoy",
            "warning"
        )
    
        return redirect(
            url_for(
                "formulario",
                token=token
            )
        )
    
    except DATABASE_ERRORS:

        if conn:
            conn.rollback()

        app.logger.exception(
            "Error en la base de datos al registrar asistencia"
        )

        flash(
            "Ocurrió un error interno",
            "error"
        )

        return redirect(url_for("inicio"))

    finally:

        if conn:
            conn.close()

@app.route("/confirmacion")
def confirmacion():

    nombre = session.pop("ultimo_nombre", None)
    token = session.get("ultimo_token")

    if not token:

        flash(
            "Sesión inválida",
            "error"
        )

        return redirect(url_for("inicio"))

    if not nombre:

        return redirect(
            url_for("formulario", token=token)
        )

    return render_template(
        "confirmacion.html",
        nombre=nombre,
        token=token
    )

@app.errorhandler(403)
def forbidden(error):
    return render_template("errors/403.html"), 403


@app.errorhandler(404)
def page_not_found(error):
    return render_template("errors/404.html"), 404


@app.errorhandler(500)
def internal_server_error(error):

    return render_template(
        "errors/500.html"
    ), 500
    
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)